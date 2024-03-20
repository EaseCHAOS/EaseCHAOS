import regex as re
import pandas as pd

import openpyxl


def _get_time_row(df: pd.DataFrame) -> pd.Series:
    """
    Get the time row from the dataframe.

    Parameters
    ----------
    df : pandas.DataFrame
        The dataframe to get the time row from.

    Returns
    -------
    pandas.Series
        The time row from the dataframe.
    """
    for row in df.iterrows():
        if re.match(r"^\d{1,2}:\d{1,2}-\d{1,2}:\d{1,2}$", str(row[1].iloc[1])):
            return row


def _get_daily_table(df: pd.DataFrame, class_pattern: str) -> pd.DataFrame:
    """
    Get the a simplified dataframe of the classes for a given class.

    Parameters
    ----------
    df : pandas.DataFrame
        The dataframe to get the simplified time table from.
        It's a general time table on a single day for all classes.
    class_pattern : str
        The class to search for. E.g. 'EL 3'
o
    Returns
    -------
    pandas.DataFrame
        The simplified dataframe for only the given class.
    """
    df = df.copy()

    time_row = _get_time_row(df)
    new_cols = time_row[1].to_list()
    new_cols.pop(0)
    new_cols.insert(0, "Classroom")
    df.columns = new_cols

    df.set_index("Classroom", inplace=True)

    df = df.iloc[time_row[0]+1:]

    df = df.mask(~df.map(lambda x: bool(re.search(class_pattern, str(x)))))
    df = df.dropna(how='all')

    return df


def _get_all_daily_tables(filename: str, class_pattern: str) -> dict:
    """
    Get all the daily tables from an excel file.

    Parameters
    ----------
    filename : str
        The filename of the excel file to get the daily tables from.
    class_pattern : str
        The class to get the daily tables or. E.g. 'EL 3'

    Returns
    -------
    dict
        A dictionary of the daily tables for each class.
    """
    workbook = openpyxl.load_workbook(filename)
    dfs = {}
    for sheet in workbook.sheetnames:
        merged_cells = workbook[sheet].merged_cells.ranges
        for mc in merged_cells.copy():
            if mc.max_col - mc.min_col == 1:
                merged_value = workbook[sheet].cell(mc.min_row, mc.min_col).value
                workbook[sheet].unmerge_cells(mc.coord)
                workbook[sheet].cell(mc.min_row, mc.min_col).value = merged_value
                workbook[sheet].cell(mc.max_row, mc.max_col).value = merged_value

        data = workbook[sheet].values
        header = next(data)
        df = pd.DataFrame(data, columns=header)
        df = df.dropna(axis=1, how="all")

        dfs[sheet] = _get_daily_table(df, class_pattern)

    return dfs


def get_time_table(filename: str, class_pattern: str) -> pd.DataFrame:
    """
    Get the complete time table for a particular class for all days.

    Parameters
    ----------
    filename : str
        The filename of the excel file. This file contains every class with the days as the sheet names. 
    class_pattern : str
        The class to get the complete time table for. E.g. 'EL 3'

    Returns
    -------
    pandas.DataFrame
        The complete time table for the given class.
    """
    daily_tables = _get_all_daily_tables(filename, class_pattern)

    
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    for key, value in daily_tables.items():
        if key.title() in days:
            columns = value.columns
            break
    else:
        raise ValueError(f"No sheet found for any of the days: {days}")

    final_df = pd.DataFrame(
        columns=columns,
        index=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"],
    )

    for day, table in daily_tables.items():
        for period, classes in table.items():
            available_classes = classes.dropna()
            if available_classes.any():
                classrooms = classes[classes.notna()].index
                available_classes = [re.sub(r'\s+', ' ', c.strip()) for c in available_classes.values]
                available_classes = [f"{c} ({classrooms[i]})" for i, c in enumerate(available_classes)]
                available_classes = '\n'.join(available_classes)
                final_df.loc[day, period] = available_classes

    return final_df
